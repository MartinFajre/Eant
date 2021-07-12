from openpyxl import Workbook
from datetime import datetime

# Como escribir un excel

wb = Workbook()

sheet = wb.active

#Escribir una celda
sheet['A1'] = 42

#Agregar en la ultima celda activa 
sheet.append([1,'Valor',3])

sheet['A3'] = datetime.now()

#Método alternativo para escribir una celda, usando coordenadas.
sheet.cell(6,4,'Carmelo')

wb.save('ejemplo.xlsx')   

### Lectura de Libros Excel ###

from openpyxl import load_workbook

wb = load_workbook('tabla.xlsx')

sheet = wb['Hoja1']

#Método alternativo para capturar todas las hojas en una lista
sheets = wb.sheetnames
sheet = wb[sheets[0]]

#Caputar la columna entera y muestro todos valores

columna = sheet['A']

for valor in columna: 
    print(valor.value)

#Caputar la columna entera y muestro todos valores

fila = sheet[2]

for valor in fila:
    print(valor.value)
    
#Recorrer las filas
filas = sheet.rows
for fila in filas:
    for celda in fila:
        print(celda.value)




